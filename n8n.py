import openpyxl
import os
import base64
from io import BytesIO

# File path on the n8n server
file_path = '/data/merge_requests.xlsx'  # Adjust path as needed

# Input data from the previous Set node
input_data = [item['json'] for item in input]

# Initialize workbook
if os.path.exists(file_path):
    # Load existing workbook
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
else:
    # Create new workbook and worksheet with headers
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    headers = ['Project Name', 'MR ID', 'Title', 'Author', 'Created At', 'URL']
    worksheet.append(headers)

# Get existing MR IDs to avoid duplicates
existing_mr_ids = set()
for row in worksheet.iter_rows(min_row=2, values_only=True):
    if row[1] is not None:  # MR ID is in the second column (index 1)
        existing_mr_ids.add(row[1])

# Prepare new rows, excluding duplicates
new_rows = []
for row in input_data:
    if row['mrId'] not in existing_mr_ids:
        new_rows.append([
            row['projectName'],
            row['mrId'],
            row['title'],
            row['author'],
            row['createdAt'],
            row['url']
        ])
        existing_mr_ids.add(row['mrId'])  # Update set to prevent duplicates within this run

# Append new rows to worksheet
for row in new_rows:
    worksheet.append(row)

# Save workbook to a BytesIO buffer
buffer = BytesIO()
workbook.save(buffer)
buffer.seek(0)

# Encode buffer as base64 for n8n binary output
base64_buffer = base64.b64encode(buffer.read()).decode('utf-8')

# Output binary data for the next node
return [{'binary': {'data': base64_buffer}}]
